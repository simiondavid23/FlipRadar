"""Constructor generic de fișiere .xlsx (openpyxl), partajat de exportatoarele
per-modul (Auto Anunțuri, Imobiliare Monitor, ...).

Extras din services/radar/excel_exporter.py — partea generică (workbook, stilizare
header, evidențiere pe rând, hyperlink pe coloana URL, auto-fit lățime). Radar își
păstrează propriul build_listings_xlsx neschimbat; extracția e pentru module noi.
"""
import io
from datetime import datetime
from typing import Callable, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

# Culori de fundal per grad (aceeași paletă ca Radar; D adăugat pentru module cu grad D).
GRADE_FILLS = {
    "A": "DCFCE7",
    "B": "DBEAFE",
    "C": "FEF9C3",
    "D": "FEE2E2",
}


def fmt_dt(raw) -> str:
    """Acceptă datetime sau string ISO și returnează formatare uniformă dd.mm.yyyy HH:MM."""
    if not raw:
        return ""
    if isinstance(raw, str):
        try:
            dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except ValueError:
            return raw
    else:
        dt = raw
    try:
        return dt.strftime("%d.%m.%Y %H:%M")
    except Exception:
        return str(raw)


def build_generic_xlsx(
    columns: Sequence[str],
    rows: Sequence[Sequence],
    hyperlink_col_idx: int,
    row_fill_fn: Optional[Callable[[Sequence], Optional[str]]] = None,
    sheet_title: str = "Export",
) -> bytes:
    """Construiește un workbook cu un singur sheet și îl returnează ca bytes.

    - columns: numele coloanelor (rândul de header).
    - rows: listă de rânduri, fiecare rând = listă de valori de celulă (aliniată la columns).
    - hyperlink_col_idx: indexul (0-based) al coloanei URL — celula devine hyperlink albastru.
      Folosește un index în afara intervalului (ex. -1) pentru a dezactiva.
    - row_fill_fn: opțional, primește rândul și întoarce un cod hex de culoare (fără '#')
      sau None — folosit pentru evidențierea rândului (ex. pe grad).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Export")[:31]  # Excel: max 31 caractere / nume sheet

    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    rows_list = list(rows)
    for r_idx, row in enumerate(rows_list, start=2):
        fill_hex = row_fill_fn(row) if row_fill_fn else None
        fill = (
            PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
            if fill_hex else None
        )
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if fill:
                cell.fill = fill
            if c_idx == hyperlink_col_idx + 1 and value:
                cell.hyperlink = value
                cell.font = Font(color="2563EB", underline="single")

    # Auto-fit aproximativ: lățime = max(header, cea mai lungă valoare) din coloană, plafonat.
    for col_idx, name in enumerate(columns, start=1):
        max_len = len(str(name))
        for r_idx in range(2, len(rows_list) + 2):
            val = ws.cell(row=r_idx, column=col_idx).value
            if val is None:
                continue
            length = len(str(val))
            if length > max_len:
                max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
