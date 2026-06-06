"""Export listinguri Radar in fisier xlsx folosind openpyxl.

Doua sheets: detalii listing-uri (cu evidentiere prin culoare per scor) +
sumar agregat. Returneaza bytes — apelantul (router-ul) decide ce face cu
ei (StreamingResponse de obicei).
"""
import io
from datetime import datetime
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ROW_FILLS = {
    "A": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
    "B": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    "C": PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
}


_COLUMNS = [
    "Titlu",
    "Platformă",
    "Scor",
    "Preț cerut (RON)",
    "Preț revânzare (RON)",
    "Marjă (RON)",
    "Marjă (%)",
    "Fee Ceiling",
    "Condiție",
    "Locație",
    "Vânzător",
    "Keyword",
    "Data postării",
    "Găsit de FlipRadar",
    "Status",
    "URL",
]


def _fmt_dt(dt) -> str:
    if not dt:
        return ""
    try:
        return dt.strftime("%d.%m.%Y %H:%M")
    except Exception:
        return ""


def build_listings_xlsx(rows: Iterable[dict]) -> bytes:
    """rows = iterable de dict-uri cu cheile produse de _listing_to_dict din router."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Deal-uri Radar"

    for col_idx, name in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    rows_list = list(rows)
    for r_idx, item in enumerate(rows_list, start=2):
        margin_value = item.get("margin_value") or 0
        values = [
            item.get("title", ""),
            (item.get("platform") or "").upper(),
            item.get("score") or "",
            round(item.get("price") or 0, 2),
            round(item.get("resale_price") or 0, 2) if item.get("resale_price") is not None else "",
            round(margin_value, 2) if margin_value is not None else "",
            round(item.get("margin_pct") or 0, 1) if item.get("margin_pct") is not None else "",
            round(item.get("fee_ceiling") or 0, 2) if item.get("fee_ceiling") is not None else "",
            item.get("condition") or "",
            item.get("location") or "",
            item.get("seller_name") or "",
            item.get("keyword_name") or "",
            _fmt_dt_iso(item.get("listed_at")),
            _fmt_dt_iso(item.get("found_at")),
            item.get("status") or "",
            item.get("url") or "",
        ]
        fill = _ROW_FILLS.get(item.get("score"))
        for c_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            if fill:
                cell.fill = fill
            if c_idx == len(_COLUMNS) and v:
                cell.hyperlink = v
                cell.font = Font(color="2563EB", underline="single")

    # Auto-fit aprox: pune lățimi pe baza header + cea mai mare valoare din coloană
    for col_idx, name in enumerate(_COLUMNS, start=1):
        max_len = len(name)
        for r_idx in range(2, len(rows_list) + 2):
            val = ws.cell(row=r_idx, column=col_idx).value
            if val is None:
                continue
            length = len(str(val))
            if length > max_len:
                max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    # Sheet sumar
    summary = wb.create_sheet("Sumar")
    total = len(rows_list)
    by_score = {"A": 0, "B": 0, "C": 0, "D": 0}
    for item in rows_list:
        s = item.get("score")
        if s in by_score:
            by_score[s] += 1
    prices = [float(it.get("price") or 0) for it in rows_list if it.get("price")]
    margins = [float(it.get("margin_pct") or 0) for it in rows_list if it.get("margin_pct") is not None]
    avg_price = sum(prices) / len(prices) if prices else 0
    avg_margin = sum(margins) / len(margins) if margins else 0

    keyword_counts = {}
    for it in rows_list:
        n = it.get("keyword_name") or "—"
        keyword_counts[n] = keyword_counts.get(n, 0) + 1
    top_keyword = max(keyword_counts.items(), key=lambda kv: kv[1])[0] if keyword_counts else "—"

    summary_rows = [
        ("Total deal-uri exportate", total),
        ("Scor A", by_score["A"]),
        ("Scor B", by_score["B"]),
        ("Scor C", by_score["C"]),
        ("Scor D", by_score["D"]),
        ("Preț mediu găsit (RON)", round(avg_price, 2)),
        ("Marjă medie (%)", round(avg_margin, 1)),
        ("Keyword cu cele mai multe deal-uri", top_keyword),
        ("Generat la", datetime.now().strftime("%d.%m.%Y %H:%M")),
    ]
    for r_idx, (label, value) in enumerate(summary_rows, start=1):
        lcell = summary.cell(row=r_idx, column=1, value=label)
        lcell.font = Font(bold=True)
        summary.cell(row=r_idx, column=2, value=value)
    summary.column_dimensions["A"].width = 40
    summary.column_dimensions["B"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fmt_dt_iso(raw) -> str:
    """Acceptă atât datetime cât și string ISO și returnează formatare uniformă."""
    if not raw:
        return ""
    if isinstance(raw, str):
        try:
            dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except ValueError:
            return raw
    else:
        dt = raw
    try:
        return dt.strftime("%d.%m.%Y %H:%M")
    except Exception:
        return str(raw)
