import io
import csv
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from app.database import get_db
from app.models.product import Product
from app.models.price_history import PriceHistory
from app.models.watchlist import WatchlistItem
from app.models.user import User
from app.utils.auth import require_feature

router = APIRouter(prefix="/api/import-export", tags=["Import/Export"])

# All import/export endpoints share one feature flag: bulk ingest/egress.
_import_export_user = require_feature("can_use_import_export")


@router.post("/import-csv")
async def import_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(_import_export_user),
):
    """Import products from a CSV file."""
    if not file.filename.endswith((".csv", ".txt")):
        raise HTTPException(status_code=400, detail="Fisierul trebuie sa fie CSV")

    content = await file.read()
    decoded = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(decoded), delimiter=None)

    # Auto-detect delimiter
    sample = decoded[:2000]
    if "\t" in sample:
        reader = csv.DictReader(io.StringIO(decoded), delimiter="\t")
    elif ";" in sample:
        reader = csv.DictReader(io.StringIO(decoded), delimiter=";")
    else:
        reader = csv.DictReader(io.StringIO(decoded), delimiter=",")

    imported = 0
    skipped = 0
    errors = []

    for i, row in enumerate(reader):
        try:
            name = row.get("name") or row.get("Name") or row.get("nume") or row.get("Nume") or row.get("product_name", "")
            if not name:
                skipped += 1
                continue

            asin = row.get("asin") or row.get("ASIN", "")
            ean = row.get("ean") or row.get("EAN", "")
            category = row.get("category") or row.get("categorie", "")
            source = row.get("source") or row.get("sursa", "")
            source_url = row.get("source_url") or row.get("url", "")
            price_str = row.get("price") or row.get("pret") or row.get("current_price", "0")
            currency = row.get("currency") or row.get("moneda", "EUR")

            price = 0
            try:
                price = float(str(price_str).replace(",", ".").replace(" ", ""))
            except ValueError:
                price = 0

            # Check duplicate ASIN within the current user's products
            if asin:
                existing = (
                    db.query(Product)
                    .filter(Product.asin == asin, Product.user_id == current_user.id)
                    .first()
                )
                if existing:
                    skipped += 1
                    continue

            product = Product(
                user_id=current_user.id,
                name=name.strip(),
                asin=asin.strip() if asin else None,
                ean=ean.strip() if ean else None,
                category=category.strip() if category else None,
                source=source.strip() if source else None,
                source_url=source_url.strip() if source_url else None,
                current_price=price if price > 0 else None,
                currency=currency.strip(),
            )
            db.add(product)
            db.flush()

            if price > 0:
                ph = PriceHistory(product_id=product.id, price=price, currency=currency, source=source)
                db.add(ph)

            imported += 1
        except Exception as e:
            errors.append(f"Rand {i+2}: {str(e)}")
            continue

    db.commit()
    return {"imported": imported, "skipped": skipped, "errors": errors[:10]}


@router.post("/import-excel")
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(_import_export_user),
):
    """Import products from an Excel file."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Fisierul trebuie sa fie Excel (.xlsx)")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Fisierul nu contine date")

    headers = [str(h).lower().strip() if h else "" for h in rows[0]]

    def get_col(names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return -1

    name_col = get_col(["name", "nume", "product_name", "produs"])
    asin_col = get_col(["asin"])
    ean_col = get_col(["ean", "barcode"])
    category_col = get_col(["category", "categorie"])
    price_col = get_col(["price", "pret", "current_price"])
    currency_col = get_col(["currency", "moneda"])
    source_col = get_col(["source", "sursa"])
    url_col = get_col(["source_url", "url", "link"])

    imported = 0
    skipped = 0

    for row in rows[1:]:
        try:
            name = str(row[name_col]).strip() if name_col >= 0 and row[name_col] else ""
            if not name or name == "None":
                skipped += 1
                continue

            asin = str(row[asin_col]).strip() if asin_col >= 0 and row[asin_col] else None
            if asin == "None":
                asin = None

            if asin:
                existing = (
                    db.query(Product)
                    .filter(Product.asin == asin, Product.user_id == current_user.id)
                    .first()
                )
                if existing:
                    skipped += 1
                    continue

            price = 0
            if price_col >= 0 and row[price_col]:
                try:
                    price = float(str(row[price_col]).replace(",", "."))
                except ValueError:
                    pass

            product = Product(
                user_id=current_user.id,
                name=name,
                asin=asin,
                ean=str(row[ean_col]).strip() if ean_col >= 0 and row[ean_col] and str(row[ean_col]) != "None" else None,
                category=str(row[category_col]).strip() if category_col >= 0 and row[category_col] and str(row[category_col]) != "None" else None,
                source=str(row[source_col]).strip() if source_col >= 0 and row[source_col] and str(row[source_col]) != "None" else None,
                source_url=str(row[url_col]).strip() if url_col >= 0 and row[url_col] and str(row[url_col]) != "None" else None,
                current_price=price if price > 0 else None,
                currency=str(row[currency_col]).strip() if currency_col >= 0 and row[currency_col] and str(row[currency_col]) != "None" else "EUR",
            )
            db.add(product)
            imported += 1
        except Exception:
            skipped += 1
            continue

    db.commit()
    wb.close()
    return {"imported": imported, "skipped": skipped}


@router.get("/export-products")
async def export_products_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(_import_export_user),
):
    """Export the current user's products to Excel."""
    products = (
        db.query(Product)
        .filter(Product.user_id == current_user.id)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["ID", "Name", "ASIN", "EAN", "Category", "Source", "Source URL", "Price", "Currency", "Created At"])

    for p in products:
        ws.append([p.id, p.name, p.asin, p.ean, p.category, p.source, p.source_url, p.current_price, p.currency,
                    p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else ""])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=flipradar_products.xlsx"},
    )


@router.get("/export-watchlist")
async def export_watchlist_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(_import_export_user),
):
    """Export user watchlist to Excel."""
    items = (
        db.query(WatchlistItem)
        .filter(WatchlistItem.user_id == current_user.id)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Watchlist"
    ws.append(["Product ID", "Product Name", "ASIN", "Price", "Currency", "Source", "Notes", "Added At"])

    for item in items:
        product = db.query(Product).filter(Product.id == item.product_id).first()
        ws.append([
            item.product_id,
            product.name if product else "N/A",
            product.asin if product else "",
            product.current_price if product else "",
            product.currency if product else "",
            product.source if product else "",
            item.notes or "",
            item.added_at.strftime("%Y-%m-%d %H:%M") if item.added_at else "",
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=flipradar_watchlist.xlsx"},
    )


@router.get("/template")
async def download_template(
    current_user: User = Depends(_import_export_user),
):
    """Download an import template Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["name", "asin", "ean", "category", "price", "currency", "source", "source_url"])
    ws.append(["Apple AirPods Pro 2", "", "194253944140", "electronics", "999.00", "RON", "emag.ro", "https://emag.ro/..."])
    ws.append(["Samsung Galaxy S24", "", "", "electronics", "3499.00", "RON", "altex.ro", "https://altex.ro/..."])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=flipradar_template.xlsx"},
    )
