import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List
from openpyxl import load_workbook, Workbook
from app.database import get_db
from app.models.user import User
from app.models.inventory import InventoryItem
from app.schemas.inventory import InventoryItemCreate, InventoryItemUpdate, InventoryItemResponse
from app.utils.auth import get_current_user
from app.services.currency_service import convert

router = APIRouter(prefix="/api/inventory", tags=["Inventory"])


@router.get("/", response_model=List[InventoryItemResponse])
def get_inventory(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List all inventory items for the current user."""
    return (
        db.query(InventoryItem)
        .filter(InventoryItem.user_id == current_user.id)
        .order_by(InventoryItem.created_at.desc())
        .all()
    )


@router.get("/stats")
def get_inventory_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Summary: item count, total units, total value (EUR reference)."""
    rows = (
        db.query(
            InventoryItem.currency,
            func.coalesce(func.sum(InventoryItem.purchase_price * InventoryItem.quantity), 0.0),
            func.coalesce(func.sum(InventoryItem.quantity), 0),
            func.count(InventoryItem.id),
        )
        .filter(InventoryItem.user_id == current_user.id)
        .group_by(InventoryItem.currency)
        .all()
    )

    total_value_eur = 0.0
    total_units = 0
    total_items = 0
    for currency, subtotal, units, items_count in rows:
        total_value_eur += convert(float(subtotal or 0), currency or "EUR", "EUR")
        total_units += int(units or 0)
        total_items += int(items_count or 0)

    return {
        "total_items": total_items,
        "total_units": total_units,
        "total_value_eur": round(total_value_eur, 2),
    }


@router.post("/", response_model=InventoryItemResponse)
def create_inventory_item(
    data: InventoryItemCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Add a new item to the inventory."""
    item = InventoryItem(user_id=current_user.id, **data.model_dump(exclude_none=True))
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.put("/{item_id}", response_model=InventoryItemResponse)
def update_inventory_item(
    item_id: int,
    data: InventoryItemUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Edit an inventory item."""
    item = (
        db.query(InventoryItem)
        .filter(InventoryItem.id == item_id, InventoryItem.user_id == current_user.id)
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="Produs inexistent in inventar")

    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(item, key, value)
    db.commit()
    db.refresh(item)
    return item


@router.delete("/{item_id}")
def delete_inventory_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete an inventory item."""
    item = (
        db.query(InventoryItem)
        .filter(InventoryItem.id == item_id, InventoryItem.user_id == current_user.id)
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="Produs inexistent in inventar")

    db.delete(item)
    db.commit()
    return {"message": "Produs eliminat din inventar"}


_INVENTORY_COLS = ["nume", "categorie", "sku", "cantitate", "pret_achizitie", "moneda", "sursa", "note"]


@router.get("/template")
def download_inventory_template(current_user: User = Depends(get_current_user)):
    """Returns an Excel template for inventory import with the expected columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventar"
    ws.append(_INVENTORY_COLS)
    # Exemplu pe randul 2 ca utilizatorul sa stie formatul
    ws.append(["Exemplu produs", "Electronice", "SKU-001", 5, 199.99, "RON", "altex.ro", "Note optionale"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_inventar.xlsx"},
    )


@router.post("/import-excel")
async def import_inventory_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Import inventory items from an Excel file with the inventory columns."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Fisierul trebuie sa fie Excel (.xlsx)")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Nu am putut citi fisierul: {exc}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Fisierul nu contine date sub randul de antet.")

    headers = [str(h).lower().strip() if h else "" for h in rows[0]]

    def get_col(*names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return -1

    col_name = get_col("nume", "name", "produs", "product")
    col_category = get_col("categorie", "category")
    col_sku = get_col("sku", "cod", "code")
    col_quantity = get_col("cantitate", "quantity", "qty", "stoc")
    col_price = get_col("pret_achizitie", "pret achizitie", "purchase_price", "price", "pret")
    col_currency = get_col("moneda", "currency")
    col_source = get_col("sursa", "source", "magazin")
    col_notes = get_col("note", "notes", "observatii")

    if col_name < 0 or col_price < 0:
        raise HTTPException(
            status_code=400,
            detail="Antetul trebuie sa contina cel putin coloanele 'nume' si 'pret_achizitie'.",
        )

    imported = 0
    skipped = 0
    errors: list[str] = []

    def cell(row, idx):
        if idx < 0 or idx >= len(row) or row[idx] is None:
            return None
        val = row[idx]
        if isinstance(val, str):
            val = val.strip()
            return val or None
        return val

    if len(rows) > 2000:
        raise HTTPException(status_code=400,
            detail="Fisierul depaseste limita de 2000 de randuri per import.")

    for i, row in enumerate(rows[1:], start=2):
        try:
            name = cell(row, col_name)
            if not name:
                skipped += 1
                continue
            name = str(name)

            raw_price = cell(row, col_price)
            if raw_price is None:
                skipped += 1
                errors.append(f"Randul {i}: lipseste pretul de achizitie.")
                continue
            try:
                price = float(str(raw_price).replace(",", "."))
            except (ValueError, TypeError):
                skipped += 1
                errors.append(f"Randul {i}: pret invalid '{raw_price}'.")
                continue

            qty_raw = cell(row, col_quantity)
            try:
                qty = int(qty_raw) if qty_raw is not None else 1
            except (ValueError, TypeError):
                qty = 1

            currency = cell(row, col_currency) or "RON"
            currency = str(currency).upper()
            if currency not in ("RON", "EUR", "USD"):
                currency = "RON"

            item = InventoryItem(
                user_id=current_user.id,
                name=name,
                category=str(cell(row, col_category)) if cell(row, col_category) else None,
                sku=str(cell(row, col_sku)) if cell(row, col_sku) else None,
                quantity=max(qty, 1),
                purchase_price=price,
                currency=currency,
                source=str(cell(row, col_source)) if cell(row, col_source) else None,
                notes=str(cell(row, col_notes)) if cell(row, col_notes) else None,
            )
            db.add(item)
            imported += 1
        except Exception as exc:
            skipped += 1
            errors.append(f"Randul {i}: {exc}")
            continue

    db.commit()
    wb.close()
    return {"imported": imported, "skipped": skipped, "errors": errors[:10]}
